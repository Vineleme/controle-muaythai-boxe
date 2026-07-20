import json
from pathlib import Path

from openpyxl import load_workbook

source = Path(r"C:\Users\Vinicius\OneDrive\Desktop\controle - muaythai boxe.xlsx")
target = Path(__file__).resolve().parents[1] / "data.js"

wb = load_workbook(source, data_only=False)
ws = wb["Alunos"]

students = []
for row in range(4, 101):
    name = ws.cell(row, 4).value
    if not name:
        continue
    modalidade = ws.cell(row, 1).value or ""
    horario = ws.cell(row, 2).value or ""
    turma = f"{modalidade} {horario}".strip()
    pago = ws.cell(row, 6).value or "Nao"
    cobrado = ws.cell(row, 5).value or "Nao"
    categoria = ws.cell(row, 10).value or "Mensal"
    if categoria not in {"Mensal", "Semestral", "Anual", "TotalPass"}:
        categoria = "Mensal"
    students.append(
        {
            "id": f"aluno-{row}",
            "aluno": str(name),
            "modalidade": str(modalidade),
            "horario": str(horario),
            "turma": turma,
            "cobrado": "Sim" if cobrado == "Sim" else "Nao",
            "pago": "Sim" if pago == "Sim" else "Nao",
            "valor": float(ws.cell(row, 8).value or 0),
            "forma": str(ws.cell(row, 9).value or "Pix/Outros"),
            "categoria": str(categoria),
            "observacao": str(ws.cell(row, 11).value or ws.cell(row, 10).value or "")
            if ws.cell(row, 11).value not in {"Mensal", "Semestral", "Anual", "TotalPass"}
            else "",
        }
    )

target.write_text(
    "window.INITIAL_STUDENTS = "
    + json.dumps(students, ensure_ascii=False, indent=2)
    + ";\n",
    encoding="utf-8",
)
print(target)
